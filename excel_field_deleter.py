from openpyxl import load_workbook
import sys

filename = sys.argv[1]
output_filename = "output"

# capture list of fields that will be kept from a file 
file = open('fields_to_keep.txt','r')
fields_to_keep = file.read().split("\n")
file.close()

workbook = load_workbook(filename=filename)

def print_rows(sheet):
    for row in sheet.iter_rows(max_row=1,values_only=True):
        print(row)

def delete_fields(sheet, workbook):


    col = 1
    idx_to_delete=[]
    for column in sheet.iter_cols(max_row=1,values_only=True):
        # print(column[0])
        if column[0] in fields_to_keep:
            pass
        else:
            # print("delete {0}, idx={1}".format(column[0],col))
            idx_to_delete.append(col)
        col=col+1
    # print(idx_to_delete)
    idx_to_delete.sort(reverse=True)
    for i in idx_to_delete:
        sheet.delete_cols(idx=i)
    # print_rows(sheet)

    workbook.save("{0}.xlsx".format(output_filename))



sheetnames = workbook.sheetnames # lists the workbook names
# print(sheetnames)
for sheet in sheetnames:
    delete_fields(workbook[sheet],workbook)

